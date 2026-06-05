"""
Populator safety-invariant tests against the real Rayzor ACQ Mini Model template.

Proves: (1) writes INPUT cells, (2) REFUSES formula cells, (3) operates on a copy (original
untouched), (4) structural diff unchanged, (5) PENDING EXCEL RECALC marker set, (6) reconcile
flags an un-recalced workbook.
"""
import json
import os
import sys
import tempfile

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import populator as P

TEMPLATE = (
    r"C:\Users\evana\Downloads\Shieldstone Underwriting Skill Sandbox"
    r"\Resia Rayzor Ranch Exemplar"
    r"\Resia Rayzor Ranch - Shieldstone ACQ Mini Model (5.04.26).xlsx"
)

pytestmark = pytest.mark.skipif(not os.path.exists(TEMPLATE), reason="Rayzor ACQ template not present")


def _write_spec(tmpdir):
    spec = {
        "meta": {"deal_name": "Test Deal", "slug": "testdeal", "routing": "ACQ", "template": TEMPLATE},
        "qa": {"t12_unmapped": 0, "formula_audit": []},
        "cells": [
            {"cell": "Pro Forma!B10", "value": 34000000, "type": "currency", "source": "test", "phase": 3},
            {"cell": "Pro Forma!B5", "value": 2007, "type": "year", "source": "test", "phase": 3},
            # B8 is a FORMULA (=B10/B6) -> must be REFUSED
            {"cell": "Pro Forma!B8", "value": 999999, "type": "currency", "source": "test (should refuse)", "phase": 3},
        ],
    }
    sp = os.path.join(tmpdir, "underwrite-spec.json")
    with open(sp, "w", encoding="utf-8") as f:
        json.dump(spec, f)
    return sp


def test_populate_writes_inputs_refuses_formulas():
    with tempfile.TemporaryDirectory() as tmp:
        sp = _write_spec(tmp)
        rep = P.populate(sp, TEMPLATE)
        assert "Pro Forma!B10" in rep.written
        assert "Pro Forma!B5" in rep.written
        # B8 is a formula -> refused, not written
        refused = [c for c, _ in rep.refused_formula]
        assert "Pro Forma!B8" in refused
        assert rep.ok is False  # ok is False because a refusal occurred (surfaces the bad spec cell)


def test_original_template_untouched():
    before_mtime = os.path.getmtime(TEMPLATE)
    with tempfile.TemporaryDirectory() as tmp:
        sp = _write_spec(tmp)
        rep = P.populate(sp, TEMPLATE)
        assert rep.draft_path != TEMPLATE
        assert os.path.exists(rep.draft_path)
    assert os.path.getmtime(TEMPLATE) == before_mtime  # original never written


def test_structural_diff_unchanged():
    with tempfile.TemporaryDirectory() as tmp:
        sp = _write_spec(tmp)
        rep = P.populate(sp, TEMPLATE)
        assert rep.structural_diff_ok, f"structure changed: {rep.structural_diff_detail}"


def test_written_value_persisted_and_formula_intact():
    with tempfile.TemporaryDirectory() as tmp:
        sp = _write_spec(tmp)
        rep = P.populate(sp, TEMPLATE)
        wb = openpyxl.load_workbook(rep.draft_path, data_only=False)
        ws = wb["Pro Forma"]
        assert ws["B10"].value == 34000000          # input written
        assert str(ws["B8"].value).startswith("=")  # formula preserved (=B10/B6)
        wb.close()


def test_pending_marker_and_recalc_guard():
    with tempfile.TemporaryDirectory() as tmp:
        sp = _write_spec(tmp)
        rep = P.populate(sp, TEMPLATE)
        # fresh openpyxl write has not been recalced by Excel
        assert P.is_recalced(rep.draft_path) is False


def test_reconcile_flags_unrecalced():
    with tempfile.TemporaryDirectory() as tmp:
        sp = _write_spec(tmp)
        rep = P.populate(sp, TEMPLATE)
        # openpyxl wrote B10 but B15 (=IRR) reads None until Excel recalcs -> flag
        rows = P.reconcile(rep.draft_path, {"irr": 0.2251}, metric_cells={"irr": "B15"})
        assert len(rows) == 1
        assert rows[0].flag is True       # not recalced -> flagged
        assert rows[0].excel is None


def test_deal_identity_check_passes_for_matching_template():
    # Rayzor template names itself "Resia Rayzor Ranch" -> matches expected.
    r = P.deal_identity_check(TEMPLATE, "Resia Rayzor Ranch", expected_units=329)
    assert r["match"] is True, r["reasons"]


def test_deal_identity_check_catches_wrong_deal():
    # Pretend the Rayzor template is being used for an "Envy Pompano Beach" underwrite
    # (the template-fork carryover scenario) -> must FAIL with a reason.
    r = P.deal_identity_check(TEMPLATE, "Envy Pompano Beach", expected_units=214)
    assert r["match"] is False
    assert any("does not match" in s or "units" in s for s in r["reasons"])


# --------------------------------------------------------------------------- #
# Wave-1 hard-gate enforcement in populate()
# --------------------------------------------------------------------------- #

def _spec_with(tmpdir, cells, meta_extra=None, qa_extra=None):
    spec = {
        "meta": {"deal_name": "Test Deal", "slug": "testdeal", "routing": "ACQ", "template": TEMPLATE},
        "qa": {"t12_unmapped": 0, "formula_audit": []},
        "cells": cells,
    }
    if meta_extra:
        spec["meta"].update(meta_extra)
    if qa_extra:
        spec["qa"].update(qa_extra)
    sp = os.path.join(tmpdir, "underwrite-spec.json")
    with open(sp, "w", encoding="utf-8") as f:
        json.dump(spec, f)
    return sp


def test_populate_refuses_sentinel_fee():
    # BL-03: B45 = 0.05 on an ACQ deal must be refused (unread Esplanade carryover).
    with tempfile.TemporaryDirectory() as tmp:
        sp = _spec_with(tmp, [
            {"cell": "Pro Forma!B45", "value": 0.05, "type": "percent", "source": "template default", "phase": 3},
        ])
        rep = P.populate(sp, TEMPLATE)
        refused = [c for c, _ in rep.refused_fee_bounds]
        assert "Pro Forma!B45" in refused
        assert rep.ok is False
        assert "Pro Forma!B45" not in rep.written


def test_populate_allows_fee_with_override_note():
    with tempfile.TemporaryDirectory() as tmp:
        sp = _spec_with(tmp, [
            {"cell": "Pro Forma!B45", "value": 0.05, "type": "percent",
             "source": "OVERRIDE: EFB structure, 5% intentional", "phase": 3},
        ])
        rep = P.populate(sp, TEMPLATE)
        assert not rep.refused_fee_bounds
        assert "Pro Forma!B45" in rep.written


def test_populate_allows_in_bounds_fee():
    with tempfile.TemporaryDirectory() as tmp:
        sp = _spec_with(tmp, [
            {"cell": "Pro Forma!B45", "value": 0.0075, "type": "percent", "source": "ACQ $25-50M band", "phase": 3},
        ])
        rep = P.populate(sp, TEMPLATE)
        assert not rep.refused_fee_bounds
        assert "Pro Forma!B45" in rep.written


def test_populate_refuses_unit_mix_when_count_blocked():
    # BL-01: a blocked unit count refuses the S3:S21 unit-mix inputs that feed S22.
    with tempfile.TemporaryDirectory() as tmp:
        sp = _spec_with(
            tmp,
            [{"cell": "Pro Forma!S10", "value": 244, "type": "integer", "source": "raw parse", "phase": 2}],
            qa_extra={"unit_count": {"blocked": True, "counted": 244, "reasons": ["marina segment"]}},
        )
        rep = P.populate(sp, TEMPLATE)
        assert "Pro Forma!S10" in rep.refused_unit_count
        assert "Pro Forma!S10" not in rep.written
        assert rep.ok is False


def test_populate_blocks_on_identity_mismatch():
    # BL-02: a recorded identity mismatch refuses populate entirely (no wrong-deal workbook).
    with tempfile.TemporaryDirectory() as tmp:
        sp = _spec_with(
            tmp,
            [{"cell": "Pro Forma!B10", "value": 75000000, "type": "currency", "source": "OM", "phase": 3}],
            meta_extra={"deal_identity": {"match": False, "reasons": ["B2 'Aviara' != Envy"]}},
        )
        rep = P.populate(sp, TEMPLATE)
        assert rep.identity_blocked is True
        assert rep.ok is False
        assert rep.written == []


def test_populate_patch_log_records_verdicts():
    # BL-07: every audited cell produces a patch_log line; an applied patch is logged.
    with tempfile.TemporaryDirectory() as tmp:
        sp = _spec_with(
            tmp,
            [{"cell": "Pro Forma!B10", "value": 75000000, "type": "currency", "source": "OM", "phase": 3}],
            qa_extra={"formula_audit": [
                {"cell": "Pro Forma!S40", "status": "bug", "expected": "=U36*12",
                 "actual": "=U36", "patch": "=U36*12", "applied": True},
                {"cell": "Pro Forma!B66", "status": "ok"},
            ]},
        )
        rep = P.populate(sp, TEMPLATE)
        joined = " | ".join(rep.patch_log)
        assert "S40" in joined and "PATCH applied" in joined
        assert "B66" in joined and "PASS" in joined
        assert "Pro Forma!S40" in rep.patches_applied


if __name__ == "__main__":
    with tempfile.TemporaryDirectory() as tmp:
        sp = _write_spec(tmp)
        rep = P.populate(sp, TEMPLATE)
        print("written:", rep.written)
        print("refused (formula):", rep.refused_formula)
        print("structural diff ok:", rep.structural_diff_ok)
        print("draft:", rep.draft_path)
        print("identity (Envy vs Rayzor template):", P.deal_identity_check(TEMPLATE, "Envy Pompano Beach", 214))
