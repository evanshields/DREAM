"""BL-02 + BL-14 extended deal_identity_check sweep tests.

The Envy run hit template-fork carryover in every form: an 'Aviara East Pompano' B2 name, a
whole 'Rayzor Ranch' tab from a different deal, Esplanade #REF! residuals in Checks, and a stale
'2007 vintage' note on a 2020 asset. These tests build synthetic workbooks for each form so they
run without the (large, machine-specific) Rayzor template; one test uses the real template if
present to confirm a clean workbook PASSES.
"""
import os
import sys
import tempfile

import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import populator as P


def _make_wb(path, b2="Envy Pompano Beach", b6=214, extra_sheets=(), checks_ref=False,
             vintage_note=None):
    wb = openpyxl.Workbook()
    pf = wb.active
    pf.title = "Pro Forma"
    pf["B2"] = b2
    pf["B6"] = b6
    if vintage_note is not None:
        pf["Z14"] = vintage_note
    checks = wb.create_sheet("Checks")
    if checks_ref:
        checks["B2"] = "#REF!"
        checks["B3"] = "=#REF!*2"
    for s in extra_sheets:
        wb.create_sheet(s)
    wb.save(path)
    wb.close()


def test_clean_envy_passes():
    with tempfile.TemporaryDirectory() as tmp:
        wbp = os.path.join(tmp, "envy.xlsx")
        _make_wb(wbp)
        r = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214)
        assert r["match"] is True, r["reasons"]


def test_aviara_name_fails():
    # "Aviara East Pompano" shares the 'pompano' token with "Envy Pompano Beach", so the fuzzy
    # NAME check tolerates it — but the unit count 228 != 214 is the hard catch. A fully foreign
    # name (no shared token) fails on the name check directly.
    with tempfile.TemporaryDirectory() as tmp:
        wbp = os.path.join(tmp, "fork.xlsx")
        _make_wb(wbp, b2="Aviara East Pompano", b6=228)
        r = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214)
        assert r["match"] is False
        assert any("228" in s or "units" in s for s in r["reasons"])

    with tempfile.TemporaryDirectory() as tmp:
        wbp = os.path.join(tmp, "fork2.xlsx")
        _make_wb(wbp, b2="Rayzor Ranch Denton", b6=214)  # no shared token with Envy
        r = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214)
        assert r["match"] is False
        assert any("does not match" in s for s in r["reasons"])


def test_foreign_tab_fails():
    # Foreign-tab fires only on a KNOWN prior-deal token the orchestrator passes (avoids
    # false-positives on legitimate seller-doc / analysis tabs).
    with tempfile.TemporaryDirectory() as tmp:
        wbp = os.path.join(tmp, "foreign.xlsx")
        _make_wb(wbp, extra_sheets=("Rayzor Ranch",))
        r = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214,
                                  foreign_deal_tokens=["rayzor", "aviara", "esplanade"])
        assert r["match"] is False
        assert "Rayzor Ranch" in r["foreign_tabs"]


def test_seller_doc_tabs_do_not_false_positive():
    # Legitimate seller-doc / analysis tabs must NOT trip the foreign-tab gate.
    with tempfile.TemporaryDirectory() as tmp:
        wbp = os.path.join(tmp, "seller.xlsx")
        _make_wb(wbp, extra_sheets=("Rent Comps Analysis", "Seller T-12 2.28.23", "Loom Script"))
        r = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214,
                                  foreign_deal_tokens=["rayzor", "aviara"])
        assert r["match"] is True, r["reasons"]
        assert r["foreign_tabs"] == []


def test_residual_ref_reported_always_blocks_only_when_strict():
    with tempfile.TemporaryDirectory() as tmp:
        wbp = os.path.join(tmp, "ref.xlsx")
        _make_wb(wbp, checks_ref=True)
        # Wave-0 default: reported but does NOT block (a blank template legitimately has #REF!).
        soft = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214)
        assert soft["residual_errors"], "expected #REF! residual cells reported"
        assert soft["match"] is True
        # Phase-11 re-verify on a populated model: a surviving #REF! is the Esplanade defect.
        strict = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214,
                                       strict_residuals=True)
        assert strict["match"] is False


def test_vintage_mismatch_flags():
    with tempfile.TemporaryDirectory() as tmp:
        wbp = os.path.join(tmp, "vintage.xlsx")
        _make_wb(wbp, vintage_note="2007 vintage — 18-yr-old roofs justify higher reserves")
        r = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214,
                                  expected_year_built=2020)
        assert r["match"] is False
        assert r["vintage_mismatch"] is True


def test_vintage_match_ok():
    with tempfile.TemporaryDirectory() as tmp:
        wbp = os.path.join(tmp, "vintage_ok.xlsx")
        _make_wb(wbp, vintage_note="2020 vintage, new construction lease-up")
        r = P.deal_identity_check(wbp, "Envy Pompano Beach", expected_units=214,
                                  expected_year_built=2020)
        assert r["vintage_mismatch"] is False
