"""
Dream Fast-Path — Mini Model populator + reconciliation gate.

Wave 3 of the fast path. Takes an underwrite-spec.json and a Mini Model template, writes the
spec's INPUT cells into a COPY of the template (never the original, never a FORMULA cell),
runs a before/after structural diff, and flags the workbook PENDING EXCEL RECALC. After a
human opens the draft in Excel once (native recalc — also the file capital partners need),
`reconcile()` re-reads with data_only=True and diffs the Excel headline values against the
Python engine's headline_metrics at tiered tolerance.

Safety invariants (Universal Rules 1, 2, 8):
  - Operate on a copy. The original template is never mutated.
  - Write only cells whose current content is NOT a formula. A spec cell targeting a formula
    cell is REFUSED and reported, not written.
  - Run the formula audit (from spec.qa.formula_audit) before writing; apply patches only if
    marked applied=true (i.e., user-approved upstream).
  - Structural diff (sheet names, merged ranges, defined names, dimensions) must be unchanged.
  - Leave a "PENDING EXCEL RECALC" marker in the Claude Log until reconcile() confirms recalc.

This module does the mechanical write + check. It does NOT decide values (the calc engine
does) and does NOT approve anything (the human gate does).
"""
from __future__ import annotations

import json
import os
import re
import shutil
import sys
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl

# The calc engine holds the stateless Wave-1 validators (fee bounds, unit-count
# reconcile, formula sweep). Import them so populate() can enforce at write time.
# acq_engine lives one dir up under engine/.
_ENGINE_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "engine")
if _ENGINE_DIR not in sys.path:
    sys.path.insert(0, _ENGINE_DIR)
try:
    import acq_engine as _ax  # noqa: E402
except Exception:  # pragma: no cover - engine import is best-effort at module load
    _ax = None

DEFAULT_SHEET = "Pro Forma"
PENDING_MARKER = "PENDING EXCEL RECALC"

# Standard Mini Model tab names. A sheet whose name carries proper-noun tokens NOT
# in this set and NOT matching the expected deal is a foreign-tab carryover signal
# (e.g. a "Rayzor Ranch" tab surviving in an Envy workbook).
STANDARD_TABS = {
    "claude log", "model inputs", "pro forma", "t-12 inputs", "t12 inputs",
    "seller t-12", "rr inputs", "rent roll inputs", "comps", "uw snapshot",
    "checks", "unit mix", "operating expenses", "sources & uses", "sources and uses",
    "amortization", "debt", "assumptions", "cover", "instructions",
}


# ---------------------------------------------------------------------------
# Results
# ---------------------------------------------------------------------------

@dataclass
class WriteReport:
    written: List[str] = field(default_factory=list)
    refused_formula: List[Tuple[str, str]] = field(default_factory=list)  # (cell, existing formula)
    refused_fee_bounds: List[Tuple[str, str]] = field(default_factory=list)  # (cell, reason) — BL-03
    refused_unit_count: List[str] = field(default_factory=list)  # S-cells refused on a blocked count — BL-01
    refused_ltv: List[str] = field(default_factory=list)        # B51/B67 refused on ltv_gate fail — BL-15
    refused_rubs: List[str] = field(default_factory=list)       # S54/U54 refused on rubs_sign fail — BL-16
    refused_exit_cap: List[str] = field(default_factory=list)   # B79 refused on exit_cap_gate fail — BL-10
    missing_cells: List[str] = field(default_factory=list)
    patches_applied: List[str] = field(default_factory=list)
    patch_log: List[str] = field(default_factory=list)  # printed formula-audit verdicts — BL-07
    identity_blocked: bool = False  # BL-02: populate refused — deal_identity mismatch in spec
    draft_path: str = ""
    structural_diff_ok: bool = True
    structural_diff_detail: Dict[str, object] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        return (
            self.structural_diff_ok
            and not self.refused_formula
            and not self.refused_fee_bounds
            and not self.refused_unit_count
            and not self.refused_ltv
            and not self.refused_rubs
            and not self.refused_exit_cap
            and not self.identity_blocked
        )


@dataclass
class ReconcileRow:
    metric: str
    python: float
    excel: Optional[float]
    delta_pct: Optional[float]
    tolerance: float
    flag: bool


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _split_addr(addr: str) -> Tuple[str, str]:
    """'Pro Forma!B10' -> ('Pro Forma','B10'); 'B10' -> (DEFAULT_SHEET,'B10')."""
    if "!" in addr:
        sheet, cell = addr.split("!", 1)
        return sheet.strip().strip("'"), cell.strip()
    return DEFAULT_SHEET, addr.strip()


def _is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


# B45 is the ACQ acquisition-fee cell on the Pro Forma tab (field-mapping-acq.md).
def _is_acq_fee_cell(sheet: str, cell: str) -> bool:
    return sheet == DEFAULT_SHEET and cell.upper() == "B45"


# S3:S21 are the per-bedroom unit-count inputs that SUM into S22 (=B6). The unit-count
# gate guards these, never B6 directly (B6 is a formula and already refused).
_UNIT_MIX_RE = re.compile(r"^S([0-9]+)$", re.IGNORECASE)


def _is_unit_mix_cell(sheet: str, cell: str) -> bool:
    if sheet != DEFAULT_SHEET:
        return False
    m = _UNIT_MIX_RE.match(cell.strip())
    return bool(m) and 3 <= int(m.group(1)) <= 21


# BL-15: B51 (LTV input) + B67 (refi loan input) feed the B52/B66 LTV formulas; refuse them
# when the LTV gate fails (B52 itself is a formula and already refused).
def _is_ltv_input_cell(sheet: str, cell: str) -> bool:
    return sheet == DEFAULT_SHEET and cell.upper() in ("B51", "B67")


# BL-16: S54 (T-12 total) + U54 (per-unit) are the Utility Reimbursement contra inputs.
def _is_rubs_cell(sheet: str, cell: str) -> bool:
    return sheet == DEFAULT_SHEET and cell.upper() in ("S54", "U54")


# BL-10: B79 is the exit-cap INPUT cell (take-the-HIGHEST).
def _is_exit_cap_cell(sheet: str, cell: str) -> bool:
    return sheet == DEFAULT_SHEET and cell.upper() == "B79"


def _has_override_note(spec_cell: Dict) -> bool:
    """An explicit human override that keeps a 5% fee on an ACQ deal must be DELIBERATE.
    Honor it only when the cell's source citation says so in words."""
    src = str(spec_cell.get("source", "")).lower()
    return "override" in src or "intentional" in src or "confirmed 5%" in src


_ERROR_TOKENS = ("#REF!", "#NUM!", "#DIV/0!", "#VALUE!", "#NAME?")


def _is_error_literal(v: str) -> bool:
    """True if the cell value is an actual Excel error (a bare error token or a formula carrying
    one) — NOT a prose sentence that merely mentions '#REF!' (e.g. a Claude Log note saying
    'Fixed all #REF! errors'). The discriminator: the value is the error itself, or a formula."""
    s = v.strip()
    if s in _ERROR_TOKENS:
        return True
    if s.startswith("=") and any(tok in s for tok in _ERROR_TOKENS):
        return True
    return False


def _structural_fingerprint(wb) -> Dict[str, object]:
    return {
        "sheets": list(wb.sheetnames),
        "defined_names": sorted(list(wb.defined_names.keys())) if hasattr(wb.defined_names, "keys") else [],
        "dims": {ws.title: f"{ws.max_row}x{ws.max_column}" for ws in wb.worksheets},
        "merged": {ws.title: sorted(str(r) for r in ws.merged_cells.ranges) for ws in wb.worksheets},
    }


# ---------------------------------------------------------------------------
# Populate
# ---------------------------------------------------------------------------

def populate(spec_path: str, template_path: str, out_path: Optional[str] = None) -> WriteReport:
    """Write spec.cells[] INPUT values into a COPY of the template.

    Returns a WriteReport. Refuses to write any cell currently containing a formula.
    """
    with open(spec_path, "r", encoding="utf-8") as f:
        spec = json.load(f)

    meta = spec.get("meta", {})
    qa = spec.get("qa", {})
    routing = str(meta.get("routing", "ACQ")).upper()
    slug = meta.get("slug", "deal")
    if out_path is None:
        base = os.path.dirname(os.path.abspath(spec_path))
        out_path = os.path.join(base, f"{slug}-DREAM-draft.xlsx")

    # 0. HARD GATE — deal identity (BL-02). If Wave 0 recorded an identity mismatch in the
    # spec, refuse to populate at all. Do not write a wrong-deal workbook. The orchestrator
    # must repoint every carryover cell and re-run the identity check before retrying.
    identity = meta.get("deal_identity")
    if isinstance(identity, dict) and identity.get("match") is False:
        report = WriteReport(draft_path="", identity_blocked=True)
        report.patch_log.append(
            "IDENTITY BLOCK: meta.deal_identity.match == False — "
            + "; ".join(identity.get("reasons", [])[:5]))
        return report

    # 1. Copy the template (never touch the original)
    shutil.copy2(template_path, out_path)

    # 2. Structural fingerprint BEFORE
    wb_before = openpyxl.load_workbook(out_path, data_only=False)
    before_fp = _structural_fingerprint(wb_before)
    wb_before.close()

    # 3. Open for writing (keep formulas)
    wb = openpyxl.load_workbook(out_path, data_only=False)
    report = WriteReport(draft_path=out_path)

    # 3a. Formula audit patches (only those marked applied — printed-patch + human-confirm gate).
    # BL-07: emit a named PASS/PATCH line per audited cell so CP-1 always shows the verdict,
    # then apply ONLY patches the orchestrator marked applied=true.
    for audit in qa.get("formula_audit", []):
        cell_name = audit.get("cell", "?")
        if audit.get("status") == "bug" and audit.get("applied") and audit.get("patch"):
            sheet, cell = _split_addr(cell_name)
            if sheet in wb.sheetnames:
                wb[sheet][cell] = audit["patch"]
                report.patches_applied.append(cell_name)
                report.patch_log.append(f"{cell_name}: PATCH applied -> {audit['patch']}")
        elif audit.get("status") == "bug":
            report.patch_log.append(
                f"{cell_name}: PATCH pending (not applied — needs human confirm)"
                + (f" -> {audit['patch']}" if audit.get("patch") else ""))
        else:
            report.patch_log.append(f"{cell_name}: PASS")

    # BL-01: a blocked unit count means refuse to write the S-column unit-mix inputs that feed
    # S22 (B6 = =S22 is already refused as a formula). Guard against shipping a 244-unit parse.
    unit_count_blocked = bool(qa.get("unit_count", {}).get("blocked"))

    # Wave-2 gate enforcement (BL-15/16/10): the engine computes these gates in synthesis and
    # records them in the spec; a failing gate (ok==false) refuses the cells it guards. An ABSENT
    # gate (not computed) is treated as ok — these refuse only on an explicit ok==false.
    def _gate_failed(name: str) -> bool:
        g = qa.get(name)
        return isinstance(g, dict) and g.get("ok") is False

    ltv_blocked = _gate_failed("ltv_gate")          # BL-15: refuse B51/B67
    rubs_blocked = _gate_failed("rubs_sign")        # BL-16: refuse S54/U54
    exit_cap_blocked = _gate_failed("exit_cap_gate")  # BL-10: refuse B79

    # 3b. Write INPUT cells
    for c in spec.get("cells", []):
        addr = c["cell"]
        sheet, cell = _split_addr(addr)
        if sheet not in wb.sheetnames:
            report.missing_cells.append(addr)
            continue
        ws = wb[sheet]
        existing = ws[cell].value
        if _is_formula(existing):
            # Universal Rule 1: never overwrite a formula cell.
            report.refused_formula.append((addr, str(existing)))
            continue
        # BL-03 fee-bounds gate: refuse a sentinel/out-of-bounds ACQ acquisition fee at B45.
        if _is_acq_fee_cell(sheet, cell) and routing == "ACQ" and _ax is not None:
            override = bool(qa.get("fee_bounds", {}).get("override")) or _has_override_note(c)
            res = _ax.assert_fee_bounds(c["value"], routing="ACQ", override=override)
            if not res.ok:
                report.refused_fee_bounds.append((addr, res.reason))
                continue
        # BL-01 unit-count gate: refuse the unit-mix S-cells when the count is blocked.
        if unit_count_blocked and _is_unit_mix_cell(sheet, cell):
            report.refused_unit_count.append(addr)
            continue
        # BL-15 LTV gate: refuse the LTV inputs (B51/B67) when computed LTV != target or B52 is a literal.
        if ltv_blocked and _is_ltv_input_cell(sheet, cell):
            report.refused_ltv.append(addr)
            continue
        # BL-16 RUBS gate: refuse S54/U54 when the reimbursement is positive or the recovery jump is unjustified.
        if rubs_blocked and _is_rubs_cell(sheet, cell):
            report.refused_rubs.append(addr)
            continue
        # BL-10 exit-cap gate: refuse B79 when the staged exit cap is not the HIGHEST of the 3 methods.
        if exit_cap_blocked and _is_exit_cap_cell(sheet, cell):
            report.refused_exit_cap.append(addr)
            continue
        ws[cell] = c["value"]
        report.written.append(addr)

    # 3c. PENDING EXCEL RECALC marker in Claude Log
    if "Claude Log" in wb.sheetnames:
        log = wb["Claude Log"]
        # append to first empty row in column A
        r = 1
        while log.cell(row=r, column=1).value not in (None, ""):
            r += 1
        log.cell(row=r, column=1, value=PENDING_MARKER)
        log.cell(row=r, column=2, value=f"populated {len(report.written)} input cells; open in Excel to recalc")

    wb.save(out_path)
    wb.close()

    # 4. Structural fingerprint AFTER + diff
    wb_after = openpyxl.load_workbook(out_path, data_only=False)
    after_fp = _structural_fingerprint(wb_after)
    wb_after.close()
    report.structural_diff_ok = (before_fp == after_fp)
    if not report.structural_diff_ok:
        report.structural_diff_detail = {
            k: {"before": before_fp[k], "after": after_fp[k]}
            for k in before_fp if before_fp[k] != after_fp[k]
        }
    return report


# ---------------------------------------------------------------------------
# Reconcile (CP-2 gate)
# ---------------------------------------------------------------------------

# Default tiered tolerances per the plan
HEADLINE_TOL = 0.005   # NOI, DSCR, IRR, bond size, exit value
LINE_TOL = 0.02        # line items

# Where each headline metric lives in the workbook (verified cell map).
# (sheet defaults to Pro Forma). Extend as the UW Snapshot map is finalized.
METRIC_CELLS_ACQ = {
    "irr": "B15",
    "equity_multiple": "B16",
    "coc_stabilized": "B17",
    "going_in_cap": "B11",
    "projected_sale_value": "B82",
}


class IdentityMismatchError(Exception):
    """Raised when CP-2 reconcile is asked to run against a workbook whose deal identity
    does not match. The caller must escalate the fork to the parent — never silently fall
    back to a transcript comparison (the Envy run's exact failure mode)."""


def reconcile(
    draft_path: str,
    headline_metrics: Dict[str, float],
    metric_cells: Optional[Dict[str, str]] = None,
    tolerances: Optional[Dict[str, float]] = None,
    identity: Optional[Dict[str, object]] = None,
    require_identity_match: bool = True,
) -> List[ReconcileRow]:
    """Re-read the (Excel-recalced) draft with data_only=True and diff against the Python
    engine headline_metrics. Returns one ReconcileRow per metric; flag=True if outside band.

    Call this AFTER the human has opened+saved the draft in Excel (so formulas are recalced).
    If a metric cell still reads None, the workbook has not been recalced -> flag it.

    BL-05 CP-2 identity gate: when `identity` is supplied (the deal_identity_check result for
    the ground-truth workbook) and `require_identity_match` is True, a mismatch raises
    IdentityMismatchError rather than reconciling against the wrong deal. On the Envy run the
    parent handed Dream the Aviara fork as ground truth; reconcile silently degraded to a
    transcript comparison and the safety gate fired without teeth. Escalate, do not degrade.
    """
    if require_identity_match and isinstance(identity, dict) and identity.get("match") is False:
        raise IdentityMismatchError(
            "CP-2 reconcile refused — ground-truth workbook failed deal_identity_check: "
            + "; ".join(identity.get("reasons", [])[:5])
            + ". Escalate the fork to the parent; do not compare against a transcript.")
    metric_cells = metric_cells or METRIC_CELLS_ACQ
    tolerances = tolerances or {}
    wb = openpyxl.load_workbook(draft_path, data_only=True)
    rows: List[ReconcileRow] = []
    for metric, cell_addr in metric_cells.items():
        if metric not in headline_metrics:
            continue
        py = float(headline_metrics[metric])
        sheet, cell = _split_addr(cell_addr)
        xl = wb[sheet][cell].value if sheet in wb.sheetnames else None
        xl = float(xl) if isinstance(xl, (int, float)) else None
        tol = tolerances.get(metric, HEADLINE_TOL)
        if xl is None:
            rows.append(ReconcileRow(metric, py, None, None, tol, flag=True))  # not recalced
        else:
            delta = abs(py - xl) / abs(xl) if xl else abs(py)
            rows.append(ReconcileRow(metric, py, xl, delta, tol, flag=(delta > tol)))
    wb.close()
    return rows


def reconcile_self_render(
    draft_path: str,
    headline_metrics: Dict[str, float],
    metric_cells: Optional[Dict[str, str]] = None,
    tolerances: Optional[Dict[str, float]] = None,
) -> List[ReconcileRow]:
    """BL-05 self-render fallback: reconcile the engine's headline_metrics against the
    openpyxl-populated workbook the populator JUST wrote (Python-vs-its-own-render).

    Use this when no valid EXTERNAL Excel ground truth exists. The self-written draft is
    ALWAYS available, so CP-2 never has to degrade to a transcript comparison. It reuses the
    same cell map + tolerances as reconcile().

    LIMITATION (documented on purpose): this catches POPULATOR bugs (a value mis-written to a
    cell), NOT engine-LOGIC bugs — the engine values are being checked against a workbook the
    engine itself drove, so it is not an independent oracle. It is the HOTL floor, not a
    substitute for an external model. When an external model exists, prefer reconcile().

    Because a freshly openpyxl-written draft is not Excel-recalced, formula cells read None;
    self-render therefore compares only the INPUT cells the populator wrote. Pass a metric_cells
    map of input-cell addresses (e.g. purchase price, units) for a meaningful self-check.
    """
    return reconcile(
        draft_path, headline_metrics,
        metric_cells=metric_cells, tolerances=tolerances,
        identity=None, require_identity_match=False,
    )


def deal_identity_check(
    workbook_path: str,
    expected_deal_name: str,
    expected_units: Optional[int] = None,
    expected_year_built: Optional[int] = None,
    foreign_deal_tokens: Optional[List[str]] = None,
    strict_residuals: bool = False,
) -> Dict[str, object]:
    """Detect template-fork carryover: a workbook still carrying a PRIOR deal's identity.

    The Envy run exposed this in every form: Pro Forma B2 read 'Aviara East Pompano'
    (228u) and never repointed; a whole 'Rayzor Ranch' tab from a different deal rode
    along; Esplanade #REF! residuals sat in Checks B2/B3; a stale '2007 vintage' capex
    note survived on a 2020 asset. The check is the HARD gate at Wave 0 (before any agent
    trusts the workbook / before populate) and re-verified at Phase 11. A False match must
    BLOCK loudly, not silently proceed.

    Always-on hard signals (flip match=False on their own — low false-positive):
      1. Pro Forma B2 asset name vs expected (fuzzy token match — tolerates legitimate renames).
      2. Pro Forma B6 units vs expected.
      3. Foreign-tab: a tab whose name contains one of `foreign_deal_tokens` (the KNOWN prior
         deals the orchestrator passes, e.g. ['aviara','esplanade','rayzor'] for an Envy run).
         We do NOT flag arbitrary non-standard tabs — seller-doc / analysis tabs are normal.
      4. Vintage-note: an assumption note referencing a build year != expected_year_built.

    Context-dependent signals (REPORTED always; flip match=False only when opted in):
      5. Residual #REF!/#NUM! cells. A blank template legitimately shows #REF! in its Checks
         tab until wired, so this is reported for visibility but only blocks when
         `strict_residuals=True` (use at Phase-11 re-verify on a POPULATED model, where a
         surviving #REF! is the Esplanade-in-Checks defect — not at Wave 0 on a fresh template).

    Returns {"match": bool, "reasons": [...], "foreign_tabs": [...],
             "residual_errors": [...], "vintage_mismatch": bool}.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    reasons: List[str] = []
    foreign_tabs: List[str] = []
    residual_errors: List[str] = []
    vintage_mismatch = False
    match = True

    exp_tokens = [t for t in re.split(r"\W+", expected_deal_name.lower()) if len(t) > 3]
    foreign_deal_tokens = [t.lower() for t in (foreign_deal_tokens or [])]

    # 1+2. Pro Forma B2 name + B6 units
    pf = wb["Pro Forma"] if "Pro Forma" in wb.sheetnames else None
    if pf is not None:
        wb_name = str(pf["B2"].value or "")
        if wb_name and exp_tokens and not any(tok in wb_name.lower() for tok in exp_tokens):
            match = False
            reasons.append(
                f"Pro Forma B2 asset name '{wb_name}' does not match expected '{expected_deal_name}' "
                f"(possible template-fork carryover / unsaved wrong-deal file)")
        if expected_units is not None:
            wb_units = pf["B6"].value
            if isinstance(wb_units, (int, float)) and abs(int(wb_units) - expected_units) > 2:
                match = False
                reasons.append(f"Pro Forma units B6={int(wb_units)} != expected {expected_units}")

    # 3. Foreign-tab sweep — KNOWN prior-deal tokens only (avoids false positives on legitimate
    #    seller-doc / analysis tabs like 'Rent Comps Analysis' or 'Seller T-12 2.28.23').
    if foreign_deal_tokens:
        for sheet in wb.sheetnames:
            low = sheet.lower().strip()
            if low in STANDARD_TABS:
                continue
            hit = next((tok for tok in foreign_deal_tokens
                        if tok in low and tok not in exp_tokens), None)
            if hit:
                foreign_tabs.append(sheet)
                match = False
                reasons.append(f"foreign tab '{sheet}' carries prior-deal token '{hit}' "
                               f"(template-fork carryover)")

    # 5. Residual #REF!/#NUM! sweep — value IS an error literal/formula (not prose mentioning it).
    #    Skip the Claude Log narration sheet.
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "claude log":
            continue
        for row in ws.iter_rows(values_only=False):
            for c in row:
                v = c.value
                if isinstance(v, str) and _is_error_literal(v):
                    residual_errors.append(f"{ws.title}!{c.coordinate}")
                    if len(residual_errors) >= 50:
                        break
            if len(residual_errors) >= 50:
                break
        if len(residual_errors) >= 50:
            break
    if residual_errors and strict_residuals:
        match = False
        reasons.append(f"{len(residual_errors)} residual #REF!/#NUM! cell(s) on a populated model "
                       f"(prior-deal carryover): {', '.join(residual_errors[:8])}"
                       f"{' ...' if len(residual_errors) > 8 else ''}")

    # 4. Vintage-note check (scan text cells for a 19xx/20xx year token that mismatches)
    if expected_year_built is not None and pf is not None:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if isinstance(v, str):
                        for yr in re.findall(r"\b(19\d{2}|20\d{2})\s*(?:vintage|built|year)", v.lower()):
                            if abs(int(yr) - expected_year_built) > 1:
                                vintage_mismatch = True
                                match = False
                                reasons.append(
                                    f"assumption note references vintage {yr} but expected "
                                    f"year_built={expected_year_built} (stale-vintage carryover)")
                if vintage_mismatch:
                    break
            if vintage_mismatch:
                break

    wb.close()
    return {
        "match": match,
        "reasons": reasons,
        "foreign_tabs": foreign_tabs,
        "residual_errors": residual_errors,
        "vintage_mismatch": vintage_mismatch,
    }


def is_recalced(draft_path: str) -> bool:
    """Heuristic: the PENDING marker still present in Claude Log AND no recalced metric =>
    not yet recalced. Used to keep the workbook out of the 'done' state until the round-trip."""
    wb = openpyxl.load_workbook(draft_path, data_only=True)
    log = wb["Claude Log"] if "Claude Log" in wb.sheetnames else None
    pending = False
    if log is not None:
        for row in log.iter_rows(min_col=1, max_col=1, values_only=True):
            if row[0] == PENDING_MARKER:
                pending = True
                break
    wb.close()
    return not pending
