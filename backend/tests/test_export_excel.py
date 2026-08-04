"""Wave D tests — the App->Excel export + reconcile endpoints.

Exercises the endpoint WIRING (DealStore -> populator -> WriteReport JSON / draft download, and the
reconcile CP-2 path incl. the identity gate) using a SYNTHETIC minimal Mini Model built with openpyxl
— so the test needs no proprietary template. The populator's own correctness is covered by the skill
suite against the real Rayzor template; here we prove the server wraps it right.
"""
import io
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

pytest.importorskip("starlette")
pytest.importorskip("openpyxl")
import openpyxl  # noqa: E402
from fastapi import FastAPI  # noqa: E402
from starlette.testclient import TestClient  # noqa: E402

from store import get_deal_store  # noqa: E402
from routers.export_excel import router  # noqa: E402

NOW = "2026-06-06T12:00:00Z"


def _client():
    app = FastAPI()
    app.include_router(router)
    return TestClient(app)


def _mini_model_bytes(asset_name="Esplanade", units=240):
    """A minimal synthetic ACQ Mini Model: Pro Forma with B2 name, B6 units, B10 input + B11 formula,
    plus a Claude Log sheet (the populator appends its PENDING marker there)."""
    wb = openpyxl.Workbook()
    pf = wb.active
    pf.title = "Pro Forma"
    pf["B2"] = asset_name
    pf["B6"] = units
    pf["B10"] = 0          # purchase price INPUT cell (blank/zero in template)
    pf["B11"] = "=B10/B6"  # FORMULA cell — populator must REFUSE to overwrite
    pf["B45"] = 0          # acq fee INPUT
    # Claude Log shaped like a real Mini Model: a 5-col header + several pre-existing log rows.
    # The populator writes its PENDING marker into the first EMPTY row of col A (cols A+B); with
    # existing rows already establishing max_row=6 and max_col=5, that write stays WITHIN the sheet
    # dimensions, so the structural-diff guard (max_row x max_col) sees no change. (On the real
    # Rayzor template the Claude Log is already populated, so this mirrors production.)
    log = wb.create_sheet("Claude Log")
    log["A1"] = "Timestamp"; log["B1"] = "Phase"; log["C1"] = "Changed"
    log["D1"] = "Discovered"; log["E1"] = "Pending"
    for row in range(2, 6):  # rows 2-5: col A filled
        log.cell(row=row, column=1, value=f"row{row}")
    # row 6: col A is EMPTY (where the marker lands) but col E carries a value so the sheet's
    # max_row is already 6 and max_col already 5 -> the marker write changes neither dimension.
    log.cell(row=6, column=5, value="-")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _spec(slug="esplanade", routing="ACQ", deal_name="Esplanade", units=240):
    return {
        "meta": {"deal_name": deal_name, "slug": slug, "routing": routing,
                 "template": "synthetic.xlsx",
                 "deal_identity": {"match": True, "reasons": []}},
        "qa": {},
        "cells": [
            {"cell": "B10", "value": 34000000, "type": "currency", "source": "OM", "phase": 3},
            {"cell": "B45", "value": 0.0075, "type": "percent", "source": "fee band", "phase": 3},
            {"cell": "B11", "value": 0.07, "type": "percent", "source": "should be refused", "phase": 3},
        ],
        "headline_metrics": {"irr": 0.2251, "equity_multiple": 2.72},
    }


def _seed_deal(spec):
    store = get_deal_store()
    rec = store.create(spec, owner="evan@shieldstone.co", now_iso=NOW)
    return rec.deal_id


def test_export_returns_write_report_with_refusals():
    c = _client()
    deal_id = _seed_deal(_spec())
    files = {"template": ("mini.xlsx", _mini_model_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = c.post(f"/api/deals/{deal_id}/export", files=files, data={"download": "false"})
    assert r.status_code == 200
    rep = r.json()
    # B10 + B45 are INPUT cells -> written; B11 is a formula -> refused (BL non-collapsible surfacing).
    assert "B10" in rep["written"]
    assert any(ref["cell"] == "B11" for ref in rep["refusals"]["formula"])
    assert rep["written_count"] >= 2


def test_export_download_streams_xlsx():
    c = _client()
    deal_id = _seed_deal(_spec())
    files = {"template": ("mini.xlsx", _mini_model_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = c.post(f"/api/deals/{deal_id}/export", files=files, data={"download": "true"})
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    assert r.content[:2] == b"PK"  # xlsx is a zip


def test_export_blocks_on_identity_mismatch():
    c = _client()
    spec = _spec()
    spec["meta"]["deal_identity"] = {"match": False, "reasons": ["B2 carryover"]}
    deal_id = _seed_deal(spec)
    files = {"template": ("mini.xlsx", _mini_model_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = c.post(f"/api/deals/{deal_id}/export", files=files, data={"download": "false"})
    assert r.status_code == 200
    rep = r.json()
    assert rep["identity_blocked"] is True
    assert rep["ok"] is False


def test_export_rejects_non_xlsx_template():
    c = _client()
    deal_id = _seed_deal(_spec())
    files = {"template": ("notes.txt", b"hello", "text/plain")}
    r = c.post(f"/api/deals/{deal_id}/export", files=files, data={"download": "false"})
    assert r.status_code == 400


def test_export_unknown_deal_404():
    c = _client()
    files = {"template": ("mini.xlsx", _mini_model_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = c.post("/api/deals/nope/export", files=files, data={"download": "false"})
    assert r.status_code == 404


def test_reconcile_flags_unrecalced_draft():
    """A freshly openpyxl-written draft has not been recalced by Excel; the formula metric cells
    read None -> flagged not-recalced."""
    c = _client()
    deal_id = _seed_deal(_spec())
    # produce a draft via export(download), then feed it straight back (un-recalced)
    files = {"template": ("mini.xlsx", _mini_model_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    exp = c.post(f"/api/deals/{deal_id}/export", files=files, data={"download": "true"})
    assert exp.status_code == 200
    draft = {"draft": ("draft.xlsx", exp.content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = c.post(f"/api/deals/{deal_id}/reconcile", files=draft, data={"require_identity_match": "true"})
    assert r.status_code == 200
    body = r.json()
    # METRIC_CELLS_ACQ (irr=B15 etc.) aren't in our minimal sheet, so excel reads None -> not_recalced.
    assert body["not_recalced"] is True
    assert body["reconciled"] is False


def test_reconcile_identity_mismatch_409():
    c = _client()
    spec = _spec()
    spec["meta"]["deal_identity"] = {"match": False, "reasons": ["fork carryover"]}
    deal_id = _seed_deal(spec)
    files = {"template": ("mini.xlsx", _mini_model_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    exp = c.post(f"/api/deals/{deal_id}/export", files=files, data={"download": "true"})
    # export is blocked (identity), so build a standalone draft to reconcile
    draft_bytes = _mini_model_bytes()
    draft = {"draft": ("draft.xlsx", draft_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = c.post(f"/api/deals/{deal_id}/reconcile", files=draft, data={"require_identity_match": "true"})
    assert r.status_code == 409  # IdentityMismatchError surfaced, not swallowed


def test_reconcile_unknown_deal_404():
    c = _client()
    draft = {"draft": ("draft.xlsx", _mini_model_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = c.post("/api/deals/nope/reconcile", files=draft, data={"require_identity_match": "true"})
    assert r.status_code == 404
