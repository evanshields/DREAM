"""
Document Intake Service — auto-detects document type, extracts data, returns prefill map.
Accepts PDF, XLSX, CSV. Uses Kimi API for type classification when heuristics are uncertain.
"""

from __future__ import annotations

import asyncio
import io
import logging
import re
import tempfile
import os
from dataclasses import dataclass, field
from typing import Any

import pandas as pd
import openpyxl
import fitz  # pymupdf

from openai import AsyncOpenAI

from .field_mapper import (
    map_t12_to_form,
    map_rent_roll_to_form,
    map_om_to_form,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Known document types
# ---------------------------------------------------------------------------

DOC_TYPE_T12 = "T-12"
DOC_TYPE_RENT_ROLL = "Rent Roll"
DOC_TYPE_OM = "Offering Memorandum"
DOC_TYPE_COSTAR = "CoStar Report"
DOC_TYPE_UNKNOWN = "Unknown"

VALID_DOC_TYPES = {DOC_TYPE_T12, DOC_TYPE_RENT_ROLL, DOC_TYPE_OM, DOC_TYPE_COSTAR, DOC_TYPE_UNKNOWN}

# ---------------------------------------------------------------------------
# Filename heuristics — maps doc type to trigger substrings (case-insensitive)
# ---------------------------------------------------------------------------

FILENAME_HEURISTICS: dict[str, list[str]] = {
    DOC_TYPE_T12: [
        "t-12", "t12", "trailing", "t_12",
        "operating statement", "income statement",
    ],
    DOC_TYPE_RENT_ROLL: [
        "rent roll", "rentroll", "rent_roll", "rental roll",
    ],
    DOC_TYPE_OM: [
        "offering memo", "om ", " om.", "offering memorandum",
        "investment overview", "investment summary",
    ],
    DOC_TYPE_COSTAR: [
        "costar", "co-star", "sales comp", "rent comp", "market report",
    ],
}

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class IntakeResult:
    doc_type: str           # "T-12", "Rent Roll", "Offering Memorandum", "CoStar Report", "Unknown"
    confidence: float       # 0–1
    detection_method: str   # "filename_heuristic" or "ai_classification"
    extracted_fields: dict  # raw extracted values keyed by semantic name
    prefill_map: dict       # keys match frontend form field names
    warnings: list[str] = field(default_factory=list)
    raw_text: str = ""      # bounded plain-text excerpt (RAW_TEXT_CHARS) for downstream deal_docs


# ---------------------------------------------------------------------------
# Main service
# ---------------------------------------------------------------------------

class IntakeService:
    """
    Processes uploaded real estate documents and returns a prefill map for the
    EFB Underwriting Mini-App form.

    Parameters
    ----------
    kimi_client:
        An instance of ``openai.AsyncOpenAI`` pointed at the Kimi API endpoint,
        or any OpenAI-compatible async client.  Used only when filename
        heuristics are inconclusive.

    Example
    -------
    from openai import AsyncOpenAI

    kimi = AsyncOpenAI(
        api_key="...",
        base_url="https://api.moonshot.cn/v1",
    )
    svc = IntakeService(kimi_client=kimi)
    result = await svc.process_file(file_bytes, "PropertyXYZ_T-12.pdf")
    """

    KIMI_MODEL = "moonshot-v1-8k"
    TEXT_EXCERPT_CHARS = 3_000
    # Raw text carried back for the jobs pipeline's deal_docs. The analyst slice prompts truncate
    # the deal_docs JSON at ~60KB, so keep this bounded below that with headroom.
    RAW_TEXT_CHARS = 50_000

    def __init__(self, kimi_client: AsyncOpenAI) -> None:
        self._kimi = kimi_client

    # ------------------------------------------------------------------ #
    # Public                                                              #
    # ------------------------------------------------------------------ #

    async def process_file(self, file_bytes: bytes, filename: str) -> IntakeResult:
        """
        Full pipeline: detect type → extract fields → map to form.

        Parameters
        ----------
        file_bytes:
            Raw bytes of the uploaded file.
        filename:
            Original filename (used for heuristic detection and extension sniffing).
        """
        warnings: list[str] = []
        ext = _file_extension(filename)

        # Step 1 — try filename heuristics
        doc_type, heuristic_confidence = _classify_by_filename(filename)
        detection_method = "filename_heuristic"

        # Step 2 — fallback to Kimi if heuristics are inconclusive
        if heuristic_confidence < 0.7:
            try:
                text_excerpt = await asyncio.to_thread(
                    _extract_text_excerpt, file_bytes, ext, self.TEXT_EXCERPT_CHARS
                )
                if text_excerpt.strip():
                    ai_type = await self._classify_with_kimi(text_excerpt)
                    if ai_type != DOC_TYPE_UNKNOWN:
                        doc_type = ai_type
                        heuristic_confidence = 0.80
                        detection_method = "ai_classification"
                else:
                    warnings.append("Could not extract text for AI classification; no readable content.")
            except Exception as exc:
                warnings.append(f"AI classification failed ({exc}); falling back to heuristic result.")

        # Step 3 — run appropriate extractor
        extracted_fields: dict = {}
        try:
            extracted_fields = await asyncio.to_thread(
                _extract_fields, file_bytes, ext, doc_type
            )
        except Exception as exc:
            warnings.append(f"Field extraction failed: {exc}")

        # Step 4 — map to form fields
        prefill_map: dict = _build_prefill_map(doc_type, extracted_fields)

        # Step 5 — bounded raw text so the caller can pass the document content downstream
        # (deal_docs on the jobs pipeline). Best-effort: a text failure never fails the intake.
        raw_text = ""
        try:
            raw_text = await asyncio.to_thread(
                _extract_text_excerpt, file_bytes, ext, self.RAW_TEXT_CHARS
            )
        except Exception as exc:
            warnings.append(f"Raw text extraction failed: {exc}")

        return IntakeResult(
            doc_type=doc_type,
            confidence=heuristic_confidence,
            detection_method=detection_method,
            extracted_fields=extracted_fields,
            prefill_map=prefill_map,
            warnings=warnings,
            raw_text=raw_text,
        )

    # ------------------------------------------------------------------ #
    # Private — AI classification                                        #
    # ------------------------------------------------------------------ #

    async def _classify_with_kimi(self, text_excerpt: str) -> str:
        prompt = (
            "You are analyzing a commercial real estate document. "
            "Based on this excerpt, identify the document type.\n"
            "Options: T-12, Rent Roll, Offering Memorandum, CoStar Report, Other\n"
            "Reply with ONLY the document type, nothing else.\n\n"
            f"Document excerpt:\n{text_excerpt}"
        )
        try:
            response = await self._kimi.chat.completions.create(
                model=self.KIMI_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0,
                max_tokens=20,
            )
            raw = response.choices[0].message.content.strip()
            return _normalize_ai_doc_type(raw)
        except Exception as exc:
            logger.warning("Kimi classification error: %s", exc)
            return DOC_TYPE_UNKNOWN


# ---------------------------------------------------------------------------
# Internal helpers — text extraction
# ---------------------------------------------------------------------------

def _file_extension(filename: str) -> str:
    """Return lowercase extension without leading dot."""
    _, ext = os.path.splitext(filename.lower())
    return ext.lstrip(".")


def _extract_text_excerpt(file_bytes: bytes, ext: str, max_chars: int) -> str:
    """Extract a plain-text excerpt from the file for AI classification."""
    if ext == "pdf":
        return _pdf_to_text(file_bytes)[:max_chars]
    if ext in ("xlsx", "xls"):
        return _xlsx_to_text(file_bytes)[:max_chars]
    if ext == "csv":
        return _csv_to_text(file_bytes)[:max_chars]
    # Try PDF fallback for unknown extensions
    try:
        return _pdf_to_text(file_bytes)[:max_chars]
    except Exception:
        return ""


def _pdf_to_text(file_bytes: bytes) -> str:
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    parts: list[str] = []
    for page in doc:
        parts.append(page.get_text())
    doc.close()
    return "\n".join(parts)


def _xlsx_to_text(file_bytes: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    parts: list[str] = []
    for row in ws.iter_rows(values_only=True):
        row_parts = [str(cell) for cell in row if cell is not None]
        if row_parts:
            parts.append(" ".join(row_parts))
    wb.close()
    return "\n".join(parts)


def _csv_to_text(file_bytes: bytes) -> str:
    df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
    return df.to_string(index=False)


# ---------------------------------------------------------------------------
# Internal helpers — filename classification
# ---------------------------------------------------------------------------

def _classify_by_filename(filename: str) -> tuple[str, float]:
    """
    Return (doc_type, confidence) based on filename substrings.
    Confidence is 0.9 if a unique match is found, 0.5 if ambiguous, 0.0 if no match.
    """
    name_lower = filename.lower()
    matched: list[str] = []

    for doc_type, triggers in FILENAME_HEURISTICS.items():
        for trigger in triggers:
            if trigger in name_lower:
                matched.append(doc_type)
                break  # one match per doc_type is enough

    if len(matched) == 1:
        return matched[0], 0.90
    if len(matched) > 1:
        # Ambiguous — return first match with low confidence
        return matched[0], 0.50
    return DOC_TYPE_UNKNOWN, 0.0


def _normalize_ai_doc_type(raw: str) -> str:
    """Map Kimi's raw response string to a canonical doc type."""
    mapping = {
        "t-12": DOC_TYPE_T12,
        "t12": DOC_TYPE_T12,
        "t 12": DOC_TYPE_T12,
        "rent roll": DOC_TYPE_RENT_ROLL,
        "rentroll": DOC_TYPE_RENT_ROLL,
        "offering memorandum": DOC_TYPE_OM,
        "offering memo": DOC_TYPE_OM,
        "om": DOC_TYPE_OM,
        "costar report": DOC_TYPE_COSTAR,
        "costar": DOC_TYPE_COSTAR,
        "co-star": DOC_TYPE_COSTAR,
        "other": DOC_TYPE_UNKNOWN,
    }
    return mapping.get(raw.lower(), DOC_TYPE_UNKNOWN)


# ---------------------------------------------------------------------------
# Internal helpers — field extraction (dispatches by doc type + extension)
# ---------------------------------------------------------------------------

def _extract_fields(file_bytes: bytes, ext: str, doc_type: str) -> dict:
    """Dispatch to the appropriate extractor."""
    if doc_type == DOC_TYPE_T12:
        text = _extract_text_excerpt(file_bytes, ext, max_chars=50_000)
        return _extract_t12_from_text(text)

    if doc_type == DOC_TYPE_RENT_ROLL:
        if ext in ("xlsx", "xls"):
            return _extract_rent_roll_from_xlsx(file_bytes)
        # Fallback: try text extraction
        text = _extract_text_excerpt(file_bytes, ext, max_chars=50_000)
        return _extract_rent_roll_from_text(text)

    if doc_type == DOC_TYPE_OM:
        text = _extract_text_excerpt(file_bytes, ext, max_chars=50_000)
        return _extract_om_from_text(text)

    if doc_type == DOC_TYPE_COSTAR:
        text = _extract_text_excerpt(file_bytes, ext, max_chars=50_000)
        return _extract_costar_from_text(text)

    return {}


# ---------------------------------------------------------------------------
# T-12 extraction
# ---------------------------------------------------------------------------

# Dollar amount pattern: optional $, digits with optional commas, optional decimal
_DOLLAR = r"\$?\s*([\d,]+(?:\.\d+)?)"
# Percentage pattern
_PCT = r"([\d.]+)\s*%"


def _first_dollar(text: str, *labels: str) -> float | None:
    """Find the first dollar amount near any of the given label patterns."""
    for label in labels:
        pattern = rf"(?i){re.escape(label)}\b[^\n]{{0,80}}{_DOLLAR}"
        m = re.search(pattern, text)
        if m:
            try:
                return float(m.group(1).replace(",", ""))
            except ValueError:
                continue
    return None


def _first_pct(text: str, *labels: str) -> float | None:
    """Find the first percentage value near any of the given label patterns."""
    for label in labels:
        pattern = rf"(?i){re.escape(label)}\b[^\n]{{0,80}}{_PCT}"
        m = re.search(pattern, text)
        if m:
            try:
                return float(m.group(1)) / 100.0
            except ValueError:
                continue
        # Also look for dollar amounts that might represent vacancy loss
    return None


def _extract_t12_from_text(text: str) -> dict:
    """
    Extract key T-12 line items from text.
    Returns raw annual dollar values and percentages.
    All values are annual totals unless noted.
    """
    extracted: dict[str, Any] = {}

    # Gross Potential Rent
    gpr = _first_dollar(text, "gross potential rent", "gross potential revenue", "GPR", "potential rent")
    if gpr is not None:
        extracted["gpr_annual"] = gpr

    # Vacancy — try percentage first, then dollar amount
    vac_pct = _first_pct(text, "vacancy", "physical vacancy", "economic vacancy")
    if vac_pct is not None:
        extracted["vacancy_pct"] = vac_pct
    else:
        vac_dollar = _first_dollar(text, "vacancy loss", "vacancy", "loss to lease")
        if vac_dollar is not None:
            extracted["vacancy_loss_annual"] = vac_dollar
            # Derive pct if we have GPR
            if gpr:
                extracted["vacancy_pct"] = round(vac_dollar / gpr, 4)

    # Effective Gross Income
    egi = _first_dollar(text, "effective gross income", "effective gross revenue", "EGI")
    if egi is not None:
        extracted["egi_annual"] = egi

    # Concessions
    conc = _first_dollar(text, "concession", "concessions", "free rent")
    if conc is not None:
        extracted["concessions_annual"] = conc
        if gpr:
            extracted["concessions_pct"] = round(conc / gpr, 4)

    # Bad debt / credit loss
    bad_debt = _first_dollar(text, "bad debt", "credit loss", "write-off", "write off")
    if bad_debt is not None:
        extracted["bad_debt_annual"] = bad_debt

    # Other income
    other_income = _first_dollar(
        text, "other income", "ancillary income", "misc income",
        "miscellaneous income", "fee income", "laundry", "parking"
    )
    if other_income is not None:
        extracted["other_income_annual"] = other_income

    # Payroll
    payroll = _first_dollar(text, "payroll", "salary", "salaries", "labor", "wages")
    if payroll is not None:
        extracted["payroll_annual"] = payroll

    # Utilities
    utils = _first_dollar(text, "utilities", "utility", "electric", "gas", "water", "sewer")
    if utils is not None:
        extracted["utilities_annual"] = utils

    # Repairs & maintenance
    rm = _first_dollar(text, "repairs and maintenance", "repair & maintenance", "r&m",
                       "maintenance", "repairs")
    if rm is not None:
        extracted["rm_annual"] = rm

    # Turnover / make-ready
    turnover = _first_dollar(text, "turnover", "make ready", "make-ready", "unit turn")
    if turnover is not None:
        extracted["turnover_annual"] = turnover

    # Administrative / marketing
    admin = _first_dollar(text, "administrative", "admin", "marketing", "advertising")
    if admin is not None:
        extracted["admin_annual"] = admin

    # Insurance
    insurance = _first_dollar(text, "insurance", "property insurance", "liability insurance",
                               "hazard insurance")
    if insurance is not None:
        extracted["insurance_annual"] = insurance

    # Management fee — try percentage first
    mgmt_pct = _first_pct(text, "management fee", "mgmt fee", "property management fee")
    if mgmt_pct is not None:
        extracted["mgmt_fee_pct"] = mgmt_pct
    else:
        mgmt_dollar = _first_dollar(text, "management fee", "mgmt fee", "property management fee")
        if mgmt_dollar is not None:
            extracted["mgmt_fee_annual"] = mgmt_dollar

    # Property taxes
    taxes = _first_dollar(text, "property tax", "real estate tax", "real property tax", "taxes")
    if taxes is not None:
        extracted["property_tax_annual"] = taxes

    # NOI
    noi = _first_dollar(text, "net operating income", "net operating", "NOI", "total NOI")
    if noi is not None:
        extracted["noi_annual"] = noi

    # Total Operating Expenses
    total_opex = _first_dollar(
        text, "total operating expenses", "total expenses", "operating expenses total"
    )
    if total_opex is not None:
        extracted["total_opex_annual"] = total_opex

    return extracted


# ---------------------------------------------------------------------------
# Rent Roll extraction — XLSX path (primary)
# ---------------------------------------------------------------------------

# Column name aliases for fuzzy matching
_RR_COLUMN_ALIASES: dict[str, list[str]] = {
    "unit":         ["unit", "unit #", "unit no", "apt", "apartment"],
    "type":         ["type", "unit type", "bed/bath", "floorplan", "plan"],
    "sqft":         ["sqft", "sq ft", "sq. ft", "square feet", "size"],
    "beds":         ["beds", "bed", "bedrooms", "br"],
    "baths":        ["baths", "bath", "bathrooms", "ba"],
    "current_rent": ["current rent", "actual rent", "lease rent", "monthly rent", "rent"],
    "market_rent":  ["market rent", "asking rent", "quoted rent"],
    "status":       ["status", "occupancy", "occupied", "vacant"],
}


def _match_column(header: str, aliases: list[str]) -> bool:
    h = header.strip().lower()
    return any(alias in h for alias in aliases)


def _detect_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    """Find the row index (1-based) that looks like the column header row."""
    unit_aliases = _RR_COLUMN_ALIASES["unit"]
    rent_aliases = _RR_COLUMN_ALIASES["current_rent"]
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).lower() for c in row if c is not None]
        has_unit = any(any(a in c for a in unit_aliases) for c in cells)
        has_rent = any(any(a in c for a in rent_aliases) for c in cells)
        if has_unit and has_rent:
            return row_idx
    return None


def _extract_rent_roll_from_xlsx(file_bytes: bytes) -> dict:
    """
    Read rent roll XLSX, aggregate by unit type.
    Returns a dict with 'unit_type_aggregates' (list) and 'total_units'.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    extracted: dict[str, Any] = {}
    warnings: list[str] = []

    header_row_idx = _detect_header_row(ws)
    if header_row_idx is None:
        wb.close()
        extracted["warnings"] = ["Could not detect header row in rent roll XLSX."]
        return extracted

    # Read all rows into a list
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(c) if c is not None else "" for c in all_rows[header_row_idx - 1]]

    # Map semantic field → column index
    col_idx: dict[str, int | None] = {field: None for field in _RR_COLUMN_ALIASES}
    for field_name, aliases in _RR_COLUMN_ALIASES.items():
        for i, h in enumerate(headers):
            if _match_column(h, aliases):
                col_idx[field_name] = i
                break

    # Collect rows
    records: list[dict] = []
    for row in all_rows[header_row_idx:]:
        if all(c is None for c in row):
            continue
        rec: dict = {}
        for field_name, idx in col_idx.items():
            rec[field_name] = row[idx] if idx is not None else None
        records.append(rec)

    if not records:
        extracted["warnings"] = ["No data rows found after header in rent roll XLSX."]
        return extracted

    # Build a DataFrame for aggregation
    df = pd.DataFrame(records)

    def _to_float(val: Any) -> float | None:
        if val is None:
            return None
        try:
            return float(str(val).replace("$", "").replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    df["current_rent_f"] = df["current_rent"].apply(_to_float)
    df["market_rent_f"] = df["market_rent"].apply(_to_float)
    df["sqft_f"] = df["sqft"].apply(_to_float)

    # Normalize beds / baths
    def _to_int(val: Any) -> int | None:
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            return None

    df["beds_i"] = df["beds"].apply(_to_int)
    df["baths_i"] = df["baths"].apply(_to_int)

    # Determine unit type label (prefer explicit type column; fall back to Xb/Xba)
    def _unit_type_label(row: pd.Series) -> str:
        t = str(row.get("type") or "").strip()
        if t and t.lower() not in ("none", "nan", ""):
            return t
        beds = row.get("beds_i")
        baths = row.get("baths_i")
        if beds is not None:
            b_label = f"{beds}BR"
            ba_label = f"/{baths}BA" if baths is not None else ""
            return b_label + ba_label
        return "Unknown"

    df["unit_type_label"] = df.apply(_unit_type_label, axis=1)

    # Aggregate by unit type
    aggregates: list[dict] = []
    for unit_type, grp in df.groupby("unit_type_label"):
        valid_rents = grp["current_rent_f"].dropna()
        valid_market = grp["market_rent_f"].dropna()
        valid_sqft = grp["sqft_f"].dropna()
        aggregates.append({
            "type":       unit_type,
            "beds":       int(grp["beds_i"].dropna().mode().iloc[0]) if not grp["beds_i"].dropna().empty else None,
            "baths":      float(grp["baths_i"].dropna().mode().iloc[0]) if not grp["baths_i"].dropna().empty else None,
            "count":      len(grp),
            "avg_rent":   round(valid_rents.mean(), 2) if not valid_rents.empty else None,
            "avg_market_rent": round(valid_market.mean(), 2) if not valid_market.empty else None,
            "avg_sqft":   round(valid_sqft.mean(), 1) if not valid_sqft.empty else None,
        })

    # Vacancy from status column
    vacancy_rate = None
    if col_idx.get("status") is not None:
        status_vals = df["status"].astype(str).str.lower()
        vacant_mask = status_vals.str.contains(r"vacant|available|avail|v\b", regex=True)
        total = len(df)
        vacant_count = vacant_mask.sum()
        if total > 0:
            vacancy_rate = round(vacant_count / total, 4)

    extracted["unit_type_aggregates"] = aggregates
    extracted["total_units"] = len(df)
    if vacancy_rate is not None:
        extracted["vacancy_rate"] = vacancy_rate
    if warnings:
        extracted["warnings"] = warnings

    return extracted


def _extract_rent_roll_from_text(text: str) -> dict:
    """
    Minimal text-based rent roll extraction when XLSX is not available.
    Returns best-effort unit counts.
    """
    extracted: dict[str, Any] = {}

    # Try to find total unit count
    m = re.search(r"(?i)total\s+units?\s*[:\-]?\s*(\d+)", text)
    if m:
        extracted["total_units"] = int(m.group(1))

    # Look for vacancy rate
    vac_pct = _first_pct(text, "vacancy", "physical vacancy", "occupancy")
    if vac_pct is not None:
        # If labeled "occupancy" the complement is vacancy
        if re.search(r"(?i)occupancy", text):
            extracted["vacancy_rate"] = round(1 - vac_pct, 4)
        else:
            extracted["vacancy_rate"] = vac_pct

    extracted["warnings"] = ["Rent roll extracted from text only — XLSX preferred for unit mix detail."]
    return extracted


# ---------------------------------------------------------------------------
# OM extraction
# ---------------------------------------------------------------------------

def _extract_om_from_text(text: str) -> dict:
    extracted: dict[str, Any] = {}

    # Property name — first non-empty line is often the title
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    if lines:
        extracted["property_name"] = lines[0][:120]

    # Unit count
    m = re.search(r"(?i)(\d[\d,]*)\s*(?:-\s*)?unit", text)
    if m:
        extracted["total_units"] = int(m.group(1).replace(",", ""))

    # Asking / list price
    price = _first_dollar(text, "asking price", "list price", "sale price", "offered at",
                          "offering price", "total price")
    if price is not None:
        extracted["asking_price"] = price

    # Year built
    m = re.search(
        r"(?i)(?:built\s+in|year\s+built\s*[:\-]?|constructed\s+in|year\s+constructed\s*[:\-]?)\s*(\d{4})",
        text,
    )
    if m:
        yr = int(m.group(1))
        if 1900 <= yr <= 2030:
            extracted["year_built"] = yr

    # Address — look for pattern: number + street
    m = re.search(
        r"\b(\d{1,5}\s+[A-Z][a-zA-Z\s]+(?:St|Ave|Blvd|Dr|Rd|Ln|Way|Ct|Circle|Place|Pkwy)\.?)",
        text,
    )
    if m:
        extracted["address"] = m.group(1).strip()

    return extracted


# ---------------------------------------------------------------------------
# CoStar extraction (best-effort market data)
# ---------------------------------------------------------------------------

def _extract_costar_from_text(text: str) -> dict:
    extracted: dict[str, Any] = {}

    # Market vacancy
    vac = _first_pct(text, "market vacancy", "submarket vacancy", "vacancy rate", "vacancy")
    if vac is not None:
        extracted["market_vacancy_rate"] = vac

    # Average market rent
    avg_rent = _first_dollar(text, "average rent", "avg rent", "market rent", "asking rent")
    if avg_rent is not None:
        extracted["market_avg_rent"] = avg_rent

    # Cap rate
    cap = _first_pct(text, "cap rate", "capitalization rate", "going-in cap")
    if cap is not None:
        extracted["market_cap_rate"] = cap

    return extracted


# ---------------------------------------------------------------------------
# Prefill map builder
# ---------------------------------------------------------------------------

def _build_prefill_map(doc_type: str, extracted: dict) -> dict:
    """Dispatch to the appropriate field mapper."""
    if doc_type == DOC_TYPE_T12:
        unit_count = extracted.get("_unit_count")  # may be set by caller; ignored gracefully
        return map_t12_to_form(extracted, unit_count or 0)

    if doc_type == DOC_TYPE_RENT_ROLL:
        return map_rent_roll_to_form(extracted)

    if doc_type == DOC_TYPE_OM:
        return map_om_to_form(extracted)

    if doc_type == DOC_TYPE_COSTAR:
        # CoStar maps directly — used as market comps context, not direct form fields
        prefill: dict = {}
        if "market_vacancy_rate" in extracted:
            prefill["market_vacancy_rate"] = extracted["market_vacancy_rate"]
        if "market_avg_rent" in extracted:
            prefill["market_avg_rent_per_unit"] = extracted["market_avg_rent"]
        if "market_cap_rate" in extracted:
            prefill["going_in_cap_rate"] = extracted["market_cap_rate"]
        return prefill

    return {}
